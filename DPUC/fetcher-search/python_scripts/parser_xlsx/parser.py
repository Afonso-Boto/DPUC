import json

import openpyxl
from elasticsearch import Elasticsearch
from pathlib import Path

path = Path("DPUC_2021.xlsx")
xlsx_file = openpyxl.load_workbook(path)
sheet = xlsx_file.active

dpucs = dict()

print(f"num of rows: {sheet.max_row}")

current_dpucid = None
for row in sheet.iter_rows(min_row=1):
    id = row[0].value
    if id == "codUC":
        continue
    if current_dpucid is None or current_dpucid != id:

        current_dpucid = id
        dpucs[current_dpucid] = dict()

        dpucs[current_dpucid]["nome"]           = row[1].value
        dpucs[current_dpucid]["ano"]            = row[2].value
        dpucs[current_dpucid]["semestre"]       = row[3].value
        dpucs[current_dpucid]["ects"]           = row[4].value
        dpucs[current_dpucid]["horas_trabalho"] = row[5].value
        dpucs[current_dpucid]["regime_faltas"]  = row[6].value

    if row[9].value:
        new_campo = row[7].value + "_" + row[8].value
        dpucs[current_dpucid][new_campo] = row[9].value

client = Elasticsearch(
    "http://localhost:9200"
)
response = client.info()
print(response)
client.close()

lst = list()

for dpuc in dpucs:
    dicio = dpucs[dpuc]
    dicio["id"] = dpuc
    lst.append(dicio)

with open("dpuc_2021.json", "w") as jsonfile:
    json.dump(lst, jsonfile)
